from openpyxl import load_workbook

wb = load_workbook("py.xlsx")

ws = wb.worksheets[0]

print()
for row in ws.iter_rows(min_row=1):
    print(f"{row[1].value}\t{row[3].value}") 