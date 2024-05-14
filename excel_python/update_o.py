from openpyxl import load_workbook

wb = load_workbook("py.xlsx")

ws = wb["Sheet1"]

ws["A1"] = "这是真实姓名"

ws.cell(2,2).value = "更新的位置"

wb.save("py.xlsx")